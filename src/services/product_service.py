from datetime import datetime
import logging
import sqlite3
from typing import Optional,List,Tuple
import re

from openpyxl import load_workbook

from utils.time import now_vn
from models.product import Product
from database import db

MIN_QUANTITY = 5

logger = logging.getLogger(__name__)

class ProductService:
    def __init__(self):
        self.db = db
        logger.info("Product service initialized")

    def row_to_product(self, row) -> Product:
        """
        chuyển 1 dòng kết quả dict-like từ db sang đối tượng Product 
        """
        return Product(
            id= row["id"],
            code= row["code"],
            category= row["category"],
            brand= row["brand"],
            price= row["price"],
            size= row["size"],
            quantity= row["quantity"],
            name= row["name"],
            imagePath= row["imagePath"],
            QRPath= row["QRPath"],
            is_active= row["is_active"],
            description= row["description"],
            updated_at= datetime.fromisoformat(row["updated_at"]),
            created_at= datetime.fromisoformat(row["created_at"])
        )

    def add_product(self,product: Product) -> int:
        '''
        thêm 1 sản phẩm mới vào db
        
        return id của sản phẩm vừa thêm
        '''
        if not product.code or not product.code.strip():
            raise ValueError("product code cannot be empty")
        if not re.fullmatch(r'[A-Za-z0-9_-]+',product.code):
            raise ValueError("invalid product code")
        if product.quantity <= 0:
            raise ValueError("product quantity must be greater than 0")
        if product.price <= 0:
            raise ValueError("product price must be greater than 0")
        try:
            '''
            Kiem tra san pham da ton tai?
                neu co -> thong bao da ton tai
                neu khong -> them san pham vao db
            '''
            if self.get_product_by_code(product.code):
                raise ValueError(f"Sản phẩm {product.name} đã tồn tại với code sản phẩm là {product.code}.")
            else:
                product.created_at = now_vn()
                product.updated_at = now_vn()

                query = """
                INSERT INTO products
                (
                code,category,name,description,brand,price,
                size,quantity,is_active,imagePath,QRPath,
                created_at,updated_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                """
                params = (
                    product.code, product.category, product.name, product.description,
                    product.brand, product.price, product.size,
                    product.quantity, product.is_active,
                    product.imagePath, product.QRPath,
                    product.created_at, product.updated_at
                )
                product_id = self.db.execute_query(query, params, fetch = False)
                product.id = product_id
                logger.info(f"added product: {product.code}_{product.name}")
                return product_id
        except sqlite3.IntegrityError as e:
            if "UNIQUE constraint failed: products.code" in str(e):
                raise ValueError(f"the product code '{product.code}' already exists.")
            raise 
        except Exception as e:
            logger.exception(f"error adding product: {e}")
            raise 
    
    def get_product_by_code(self, code: str) -> Optional[Product]:
        '''
        tim san pham theo code:
            neu tim thay thi tra ve Product
            neu khong tim thay thi tra ve None
        '''
        try:
            query = "SELECT * FROM products WHERE code = ?"
            result = self.db.execute_query(query,(code,))
            
            if not result:
                return None
            return self.row_to_product(result[0])
        except Exception as e:
            logger.exception(f"error getting product by code: {e}")
            return None
    
    def get_product_by_id(self, id:int) -> Optional[Product]:
        '''
        tim san pham theo id:
            neu tim thay thi tra ve Product
            neu khong tim thay thi tra ve None
        '''
        try:
            query = "SELECT * FROM products WHERE id = ? AND is_active = 1"
            result = self.db.execute_query(query,(id,))

            if result:
                return self.row_to_product(result[0])
            return None
        except Exception as e:
            logger.exception(f"error getting product by id: {e}")
            return None
        
    def update_product(self,product: Product) -> bool:
        """
        Cap nhat thong tin san pham thong qua id
        return 
            true -> cap nhat thanh cong
            false -> cap nhat that bai
        """
        try:    
            if not product.id:
                raise ValueError("Ma san pham khong hop le")
            
            existing = self.get_product_by_code(product.code)
            if existing and existing.id != product.id:
                raise ValueError(f"code san pham {product.code} da duoc su dung")
            
            product.updated_at = now_vn()
            query = """
            UPDATE products
            SET code = ?, name = ?, description = ?,
                brand = ?, price = ?, size = ?,
                quantity = ?, imagePath = ?, QRPath = ?,
                updated_at = ?, category = ?
            WHERE id = ? and is_active = 1
            """
            params = (
                product.code, product.name, product.description,
                product.brand, product.price, product.size,
                product.quantity, product.imagePath, product.QRPath,
                product.updated_at, product.category, product.id
            )
            self.db.execute_query(query,params,fetch=False)
            logger.info(f"updated product: {product.code}_{product.name}")
            return True
        except Exception as e:
            logger.exception(f"error updating product: {e}")
            raise
        
    def delete_product(self, id: int) -> bool:
        """
        xoa san pham theo huong soft delete
        set is_active = 0
        """
        try:
            query = """
            UPDATE products
            SET is_active = 0
            WHERE id = ?
            """
            self.db.execute_query(query, (id,), fetch=False)

            logger.info(f"deleted product: {id}")
            return True
        except Exception as e:
            logger.exception(f"error deleting product: {e}")
            raise

    def get_all_product(self):
        """
        lay tat ca san pham con hoat dong
        """
        try:
            query = """
            SELECT * FROM products
            WHERE is_active = 1
            ORDER BY name
            """
            results = self.db.execute_query(query)
            products = []
            for row in results:
                products.append(self.row_to_product(row))

            logger.info(f"there are {len(products)} products")
            return products
        except Exception as e:
            logger.exception(f"error getting all products: {e}")
            return[]
    def search_products(): pass
    def get_low_stock_product(self, threshold = MIN_QUANTITY) -> List[Product]:
        """
        lay san pham sap het hang
        """
        try:
            query = """
            SELECT * FROM products
            WHERE quantity <= ? AND quantity > 0 AND is_active = 1
            ORDER BY quantity ASC
            """
            
            results = self.db.execute_query(query,(threshold, ))

            products = []
            for row in results:
                products.append(self.row_to_product(row))

            logger.info(f"there are {len(products)} low stock products")
            return products
        except Exception as e:
            logger.exception(f"error getting low stock products: {e}")
            return []
        
    def get_out_of_stock_product(self) -> List[Product]:
        """
        Lay san pham het hang
        """
        try:
            query = """
            SELECT * FROM products
            WHERE quantity = 0 and is_active = 1
            ORDER BY name
            """

            results = self.db.execute_query(query)
            products = []
            for row in results:
                products.append(self.row_to_product(row))
            
            logger.info(f"there are {len(products)} out of stock products")
            return products
        except Exception as e:
            logger.exception(f"error getting out of stock products: {e}")
            return []
        
    def update_product_quantity(self,product_id: int,
                                quantity_change) -> bool:
        """
        cap nhat so luong ton kho:
            khi ban hang(quanity change am)
            khi nhap hang(quantity change duong)
        """
        try:
            #kiem tra san pham co ton tai?
            product = self.get_product_by_id(product_id)
            if not product:
                raise ValueError(f"can't find product with id: {product_id}")
            
            #kiem tra so luong san pham co hop le?
            new_quantity = product.quantity + quantity_change
            if new_quantity < 0:
                raise ValueError(f"product quantity can't be negative! current quantity: {product.quantity}")
            
            #thuc hien cap nhat
            product.updated_at = now_vn()
            query = """
            UPDATE products
            SET quantity = quantity + ?
                updated_at = ?
            WHERE id = ? AND is_active = 1
            """
            params = (
                quantity_change,
                product.updated_at, 
                product_id
            )
            self.db.execute_query(query,params,fetch= False)
            
            logger.info(f"updated product quantity: {product_id} by {quantity_change}")
            return True
        except Exception as e:
            logger.exception(f"error updating product quantity: {e}")
            raise 
    
    def get_brands(self) -> List[str]:
        """
        lay danh sach cac nhan hang
        """
        try:
            query = """
           SELECT DISTINCT brand FROM products 
            WHERE brand IS NOT NULL AND brand != ''
                AND is_active = 1
            ORDER BY brand"""

            results = self.db.execute_query(query)
            brands = []
            for row in results:
                brands.append(row[0])
            return brands
        except Exception as e:
            logger.exception(f"error getting brands: {e}")
            return []
    
    def get_categories(self) -> List[str]:
        """
        lay danh sach cac loai san pham
        """
        try:
            query = """
            SELECT * FROM products
            WHERE category IS NOT NULL AND category != ''
                AND is_active = 1
            ORDER BY category"""

            results = self.db.execute_query(query)
            categories = []
            for row in results:
                categories.append(row[0])
            return categories
        except Exception as e:
            logger.exception(f"error getting categories: {e}")
            return []

    def import_from_excel(self, file_path: str, sheet_name: str = 'products'):
        """
        lấy dữ liệu từ file excel và lưu vào db
            -xử lý từng dòng
            -không rollback toàn bộ nếu 1 dòng lỗi
        Args:
            file_path (str): _description_
            sheet_name (str, optional): _description_. Defaults to 'products'.
        """
        reports = {
            "success": 0,
            "failed" : 0,
            "errors": []
        }

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            
            for idx, row in enumerate(sheet.iter_rows(min_row= 2,values_only= True), start=2):
                # bỏ qua các dòng trống
                if row and any(row): 
                    data = dict(zip(headers,row))

                    try:
                        product = Product(
                            code = data.get("code"),
                            category = data.get("category"),
                            brand = data.get("brand"),
                            price = data.get("price"),
                            size = data.get("size"),
                            quantity = data.get("quantity"),
                            name = data.get("name"),
                            imagePath = data.get("imagePath"),
                            QRPath = data.get("QRPath"),
                            is_active = 1,
                            description = data.get("description"),
                        )

                        self.add_product(product)
                        reports["success"] += 1
                    except Exception as row_error:
                        reports["failed"] += 1
                        reports["errors"].append({
                            "row" : idx,
                            "code": data.get("code"),
                            "error": str(row_error)
                        })
                        logger.warning(f"import error at row {idx}: {row_error}")
                        continue
            return reports
        except Exception as e:
            logger.exception(f"error importing from excel: {e}")
            raise
        
